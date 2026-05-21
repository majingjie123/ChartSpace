"""
Excel 多空间智能图表分析工具
基于 Flask + SQLAlchemy + Vue 3 + Plotly.js
"""
import os
import sys
import uuid
import json
import logging
from logging.handlers import TimedRotatingFileHandler
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import threading
import psutil
import argparse
from flask import (Flask, jsonify, request, send_file, send_from_directory,
                   render_template, session)
from flask_cors import CORS
from sqlalchemy import (Column, Integer, String, Float, Boolean, DateTime,
                        Text, ForeignKey, create_engine, case, inspect, text)
from sqlalchemy.orm import (DeclarativeBase, Mapped, mapped_column,
                            relationship, sessionmaker, scoped_session)

# ---------------------------------------------------------------------------
# 路径处理：兼容开发环境和 PyInstaller 打包环境
# ---------------------------------------------------------------------------
def get_app_dir():
    """获取应用数据目录（exe 所在目录或项目根目录）"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_path(relative_path):
    """获取资源文件路径（模板等内置资源）"""
    if getattr(sys, '_MEIPASS', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

APP_DIR = get_app_dir()
DB_PATH = os.path.join(APP_DIR, 'excelany.db')
UPLOAD_FOLDER = os.path.join(APP_DIR, 'uploads')
LOG_FOLDER = os.path.join(APP_DIR, 'logs')
BACKUP_FOLDER = os.path.join(APP_DIR, 'backup')

# ---------------------------------------------------------------------------
# Flask 应用初始化
# ---------------------------------------------------------------------------
app = Flask(__name__, 
            template_folder=get_resource_path('templates'),
            static_folder=get_resource_path('static'))
# 修改 Jinja2 定界符，避免与 Vue.js 的 {{ }} 冲突
app.jinja_env.variable_start_string = '(('
app.jinja_env.variable_end_string = '))'
app.jinja_env.block_start_string = '(%'
app.jinja_env.block_end_string = '%)'
app.jinja_env.comment_start_string = '(#'
app.jinja_env.comment_end_string = '#)'
CORS(app)

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'excelany-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 上传限制
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 确保目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# 全局安全配置
# ---------------------------------------------------------------------------
ACCESS_PASSWORD = None

@app.before_request
def check_auth():
    """可选启动密码校验"""
    if ACCESS_PASSWORD is None:
        return
    # 允许访问登录接口和静态资源
    if request.path in ['/api/auth/login', '/api/auth/status'] or \
       request.path.startswith('/static/') or \
       request.path == '/':
        return
    if not session.get('is_authenticated'):
        return jsonify({'error': '未授权访问', 'code': 401}), 401

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    return jsonify({
        'required': ACCESS_PASSWORD is not None,
        'authenticated': session.get('is_authenticated', False)
    })

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json(silent=True) or {}
    pwd = data.get('password')
    if ACCESS_PASSWORD and pwd == ACCESS_PASSWORD:
        session['is_authenticated'] = True
        return jsonify({'message': '登录成功'})
    return jsonify({'error': '密码错误'}), 401

# ---------------------------------------------------------------------------
# 内存监控工具
# ---------------------------------------------------------------------------
def get_memory_info():
    """获取当前进程内存占用及系统内存状态"""
    process = psutil.Process(os.getpid())
    mem_info = process.memory_info()
    sys_mem = psutil.virtual_memory()
    return {
        'process_mb': round(mem_info.rss / (1024 * 1024), 2),
        'sys_percent': sys_mem.percent,
        'sys_available_mb': round(sys_mem.available / (1024 * 1024), 2),
        'warning': sys_mem.percent > 85
    }

def categorize_data_volume(row_count):
    """数据量级判定"""
    if row_count >= 1000000:
        return "million", f"该文件包含 {row_count} 行数据，属于百万行级及以上海量表格。加载、渲染、运算耗时较长，建议开启采样以保证流畅度。"
    elif row_count >= 100000:
        return "hundred_thousand", f"该文件包含 {row_count} 行数据，属于十万行级大文件，加载预计耗时较长。"
    elif row_count >= 10000:
        return "ten_thousand", f"该文件包含 {row_count} 行数据，属于万行级数据，加载较快。"
    return "small", f"该文件包含 {row_count} 行数据，属于常规量级。"

@app.route('/api/system/memory', methods=['GET'])
def system_memory():
    return jsonify(get_memory_info())

# ---------------------------------------------------------------------------
# 日志配置（按天滚动）
# ---------------------------------------------------------------------------
log_handler = TimedRotatingFileHandler(
    os.path.join(LOG_FOLDER, 'error.log'),
    when='midnight',
    interval=1,
    backupCount=30,
    encoding='utf-8'
)
log_handler.setFormatter(logging.Formatter(
    '%(asctime)s [%(levelname)s] %(message)s'
))
log_handler.setLevel(logging.ERROR)
app.logger.addHandler(log_handler)
app.logger.setLevel(logging.ERROR)

# ---------------------------------------------------------------------------
# 数据库引擎与会话
# ---------------------------------------------------------------------------
engine = create_engine(
    app.config['SQLALCHEMY_DATABASE_URI'],
    echo=False,
    connect_args={'check_same_thread': False}
)
db_session = scoped_session(sessionmaker(bind=engine))

# ---------------------------------------------------------------------------
# SQLAlchemy ORM 基类
# ---------------------------------------------------------------------------
class Base(DeclarativeBase):
    pass

# ============================== 数据库模型 ===================================

class Space(Base):
    """空间/工作区"""
    __tablename__ = 'space'
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(100), nullable=False, default='新空间')
    created_at = Column(DateTime, default=datetime.utcnow)

    datasets = relationship('Dataset', back_populates='space', cascade='all, delete-orphan')
    charts = relationship('Chart', back_populates='space', cascade='all, delete-orphan')
    chat_histories = relationship('ChatHistory', back_populates='space', cascade='all, delete-orphan')
    notes = relationship('AnalysisNote', back_populates='space', cascade='all, delete-orphan')
    space_ai_configs = relationship('SpaceAIConfig', back_populates='space', cascade='all, delete-orphan')

    def to_dict(self):
        return {'id': self.id, 'name': self.name, 'created_at': self.created_at.isoformat() if self.created_at else None}


class Dataset(Base):
    """数据集"""
    __tablename__ = 'dataset'
    id = Column(Integer, primary_key=True, autoincrement=True)
    space_id = Column(Integer, ForeignKey('space.id', ondelete='CASCADE'), nullable=False)
    name = Column(String(100), nullable=False)
    file_path = Column(String(200), nullable=False)
    selected_sheet = Column(String(100), nullable=True)
    preprocessing_options = Column(Text, nullable=True)  # JSON
    row_count = Column(Integer, nullable=True)
    uploaded_at = Column(DateTime, default=datetime.utcnow)

    space = relationship('Space', back_populates='datasets')
    charts = relationship('Chart', back_populates='dataset', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'space_id': self.space_id, 'name': self.name,
            'file_path': self.file_path, 'selected_sheet': self.selected_sheet,
            'preprocessing_options': self.preprocessing_options,
            'row_count': self.row_count,
            'uploaded_at': self.uploaded_at.isoformat() if self.uploaded_at else None
        }


class Chart(Base):
    """图表"""
    __tablename__ = 'chart'
    id = Column(Integer, primary_key=True, autoincrement=True)
    space_id = Column(Integer, ForeignKey('space.id', ondelete='CASCADE'), nullable=False)
    dataset_id = Column(Integer, ForeignKey('dataset.id', ondelete='CASCADE'), nullable=False)
    name = Column(String(100), nullable=False)
    chart_type = Column(String(20), default='scatter')
    x_col = Column(String(50), nullable=True)
    y_col = Column(String(50), nullable=True)
    y2_col = Column(String(50), nullable=True)
    trend_enabled = Column(Boolean, default=True)
    config = Column(Text, nullable=True)  # JSON: { title, x_label, y_label, color }
    created_at = Column(DateTime, default=datetime.utcnow)

    space = relationship('Space', back_populates='charts')
    dataset = relationship('Dataset', back_populates='charts')

    def to_dict(self):
        return {
            'id': self.id, 'space_id': self.space_id, 'dataset_id': self.dataset_id,
            'name': self.name, 'chart_type': self.chart_type,
            'x_col': self.x_col, 'y_col': self.y_col, 'y2_col': self.y2_col,
            'trend_enabled': bool(self.trend_enabled), 'config': self.config,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class AIConfig(Base):
    """AI 配置"""
    __tablename__ = 'ai_config'
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(100), nullable=False)
    base_url = Column(String(200), nullable=False)
    api_key = Column(String(200), nullable=False)
    model = Column(String(100), nullable=False)
    system_prompt = Column(Text, nullable=True)
    max_tokens = Column(Integer, default=2000)
    temperature = Column(Float, default=0.7)
    is_default = Column(Boolean, default=False)
    cached_models = Column(Text, nullable=True, default='[]')

    space_ai_configs = relationship('SpaceAIConfig', back_populates='ai_config', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'name': self.name, 'base_url': self.base_url,
            'api_key': self.api_key, 'model': self.model,
            'system_prompt': self.system_prompt, 'max_tokens': self.max_tokens,
            'temperature': self.temperature, 'is_default': bool(self.is_default),
            'cached_models': json.loads(self.cached_models) if self.cached_models else []
        }


class SpaceAIConfig(Base):
    """空间 - AI 配置关联"""
    __tablename__ = 'space_ai_config'
    id = Column(Integer, primary_key=True, autoincrement=True)
    space_id = Column(Integer, ForeignKey('space.id', ondelete='CASCADE'), nullable=False)
    ai_config_id = Column(Integer, ForeignKey('ai_config.id', ondelete='CASCADE'), nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow)

    space = relationship('Space', back_populates='space_ai_configs')
    ai_config = relationship('AIConfig', back_populates='space_ai_configs')

    def to_dict(self):
        return {
            'id': self.id, 'space_id': self.space_id,
            'ai_config_id': self.ai_config_id,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class ChatHistory(Base):
    """聊天记录"""
    __tablename__ = 'chat_history'
    id = Column(Integer, primary_key=True, autoincrement=True)
    space_id = Column(Integer, ForeignKey('space.id', ondelete='CASCADE'), nullable=False)
    role = Column(String(20), nullable=False)  # 'user' or 'assistant'
    content = Column(Text, nullable=False)
    dataset_id = Column(Integer, ForeignKey('dataset.id', ondelete='SET NULL'), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    space = relationship('Space', back_populates='chat_histories')

    def to_dict(self):
        return {
            'id': self.id, 'space_id': self.space_id, 'role': self.role,
            'content': self.content, 'dataset_id': self.dataset_id,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class AnalysisNote(Base):
    """分析笔记"""
    __tablename__ = 'analysis_note'
    id = Column(Integer, primary_key=True, autoincrement=True)
    space_id = Column(Integer, ForeignKey('space.id', ondelete='CASCADE'), nullable=False)
    title = Column(String(200), nullable=True)
    content = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    space = relationship('Space', back_populates='notes')

    def to_dict(self):
        return {
            'id': self.id, 'space_id': self.space_id, 'title': self.title,
            'content': self.content,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

# ============================== 工具函数 =====================================

class APIError(Exception):
    """API 业务异常"""
    def __init__(self, message, status_code=400):
        super().__init__(message)
        self.status_code = status_code

@app.errorhandler(APIError)
def handle_api_error(error):
    return jsonify({'error': str(error)}), error.status_code

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': '接口不存在'}), 404

@app.errorhandler(405)
def method_not_allowed(error):
    valid = getattr(error, 'valid_methods', None)
    msg = f'请求方法不允许。支持的请求方法: {", ".join(valid) if valid else "N/A"}'
    return jsonify({'error': msg}), 405

@app.errorhandler(413)
def too_large(error):
    return jsonify({'error': '文件过大，最大支持 50MB'}), 413

@app.errorhandler(Exception)
def handle_generic_error(error):
    app.logger.error(f'未捕获异常: {request.path}', exc_info=True)
    return jsonify({'error': '服务器内部错误，请查看日志'}), 500

# ============================== API 路由 =====================================

# -----------------------------------------------------------------------
# 空间管理
# -----------------------------------------------------------------------
@app.route('/api/spaces', methods=['GET'])
def list_spaces():
    """获取所有空间列表"""
    spaces = db_session.query(Space).order_by(Space.created_at.desc()).all()
    return jsonify([s.to_dict() for s in spaces])


@app.route('/api/spaces', methods=['POST'])
def create_space():
    """创建空间"""
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    if not name or not name.strip():
        raise APIError('空间名称不能为空')
    space = Space(name=name.strip())
    db_session.add(space)
    db_session.commit()
    return jsonify(space.to_dict()), 201


@app.route('/api/spaces/<int:space_id>', methods=['PUT'])
def rename_space(space_id):
    """重命名空间"""
    space = db_session.get(Space, space_id)
    if not space:
        raise APIError('空间不存在', 404)
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    if not name or not name.strip():
        raise APIError('空间名称不能为空')
    space.name = name.strip()
    db_session.commit()
    return jsonify(space.to_dict())


@app.route('/api/spaces/<int:space_id>', methods=['DELETE'])
def delete_space(space_id):
    """删除空间（级联删除关联数据）"""
    space = db_session.get(Space, space_id)
    if not space:
        raise APIError('空间不存在', 404)
    db_session.delete(space)
    db_session.commit()
    return jsonify({'message': '已删除'})


# -----------------------------------------------------------------------
# 数据集管理
# -----------------------------------------------------------------------
def _get_dataframe(dataset_id):
    """根据 dataset_id 读取 DataFrame（应用已存储的预处理选项）"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    file_ext = os.path.splitext(ds.file_path)[1].lower()
    if file_ext not in ('.xlsx', '.xls'):
        raise APIError('不支持的文件格式')
    # 超大文件分块读取（>100MB 仅读取前 50000 行）
    file_size_mb = _get_file_size_mb(ds.file_path)
    if file_size_mb > 100:
        app.logger.info(f'大文件({file_size_mb:.0f}MB)采用分块读取策略')
        df = pd.read_excel(ds.file_path, sheet_name=ds.selected_sheet or 0, engine='openpyxl', nrows=50000)
    else:
        df = pd.read_excel(ds.file_path, sheet_name=ds.selected_sheet or 0, engine='openpyxl')
    # 应用已存储的预处理选项
    import json
    opts = json.loads(ds.preprocessing_options) if ds.preprocessing_options else {}
    # 缺失值处理
    missing = opts.get('missing')
    if missing == 'drop':
        df = df.dropna()
    elif missing == 'ffill':
        df = df.ffill()
    elif missing == 'linear':
        df = df.interpolate(method='linear')
    elif missing == 'mean':
        for col in df.select_dtypes(include=[np.number]).columns:
            df[col] = df[col].fillna(df[col].mean())
    elif missing == 'median':
        for col in df.select_dtypes(include=[np.number]).columns:
            df[col] = df[col].fillna(df[col].median())
    # 采样
    sampling = opts.get('sampling')
    if sampling and isinstance(sampling, dict):
        method = sampling.get('method')
        n = int(sampling.get('n', 50000))
        if method == 'random' and len(df) > n:
            df = df.sample(n=n, random_state=42)
        elif method == 'equidistant' and len(df) > n:
            step = len(df) // n
            df = df.iloc[::step]
    # 日期列转数值
    x_type = opts.get('x_type')
    x_col = opts.get('x_col')
    if x_type == 'timestamp' and x_col and x_col in df.columns:
        df[x_col + '_num'] = pd.to_datetime(df[x_col]).astype(np.int64) // 10**9
    return df


@app.route('/api/spaces/<int:space_id>/datasets', methods=['GET'])
def list_datasets(space_id):
    """获取空间内所有数据集"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    datasets = db_session.query(Dataset).filter_by(space_id=space_id).order_by(Dataset.uploaded_at.desc()).all()
    return jsonify([d.to_dict() for d in datasets])


def get_row_count(file_path, sheet_name=None):
    """高效获取 Excel 行数（不加载全部数据到内存）"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, keep_links=False)
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
        else:
            ws = wb.active
        count = ws.max_row
        wb.close()
        return count
    except Exception:
        # 备选方案：使用 pandas 但限制列以减少内存
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, usecols=[0], engine='openpyxl')
            return len(df)
        except Exception:
            return 0

@app.route('/api/spaces/<int:space_id>/datasets', methods=['POST'])
def upload_dataset(space_id):
    """上传 Excel 文件（第一步：上传并解析工作表名）"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    if 'file' not in request.files:
        raise APIError('请上传文件')
    file = request.files['file']
    if file.filename == '':
        raise APIError('请选择文件')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        raise APIError('仅支持 .xlsx 或 .xls 文件')
    # UUID 重命名保存
    safe_name = f"{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(UPLOAD_FOLDER, safe_name)
    file.save(save_path)
    # 读取工作表名
    xls = pd.ExcelFile(save_path, engine='openpyxl')
    sheet_names = xls.sheet_names
    xls.close()
    # 暂存上传记录（未确认 sheet），返回 sheets 供前端选择
    ds = Dataset(
        space_id=space_id,
        name=os.path.splitext(file.filename)[0],
        file_path=save_path,
        selected_sheet=None,
        row_count=None
    )
    db_session.add(ds)
    db_session.commit()
    
    # 尝试高效统计行数
    ds.row_count = get_row_count(save_path)
    db_session.commit()
    vol_level, vol_msg = categorize_data_volume(ds.row_count)
    
    return jsonify({
        'dataset_id': ds.id, 
        'sheets': sheet_names, 
        'name': ds.name,
        'vol_level': vol_level,
        'vol_msg': vol_msg
    })


@app.route('/api/datasets/<int:dataset_id>/sheets', methods=['GET'])
def get_dataset_sheets(dataset_id):
    """获取数据集文件的工作表名称列表"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    xls = pd.ExcelFile(ds.file_path, engine='openpyxl')
    sheets = xls.sheet_names
    xls.close()
    return jsonify(sheets)


@app.route('/api/datasets/<int:dataset_id>/confirm-sheet', methods=['POST'])
def confirm_dataset_sheet(dataset_id):
    """确认选择的工作表，完成数据集创建"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    data = request.get_json(silent=True) or {}
    sheet = data.get('sheet', '')
    if sheet:
        ds.selected_sheet = sheet
    # 高效统计行数
    ds.row_count = get_row_count(ds.file_path, sheet)
    vol_level, vol_msg = categorize_data_volume(ds.row_count or 0)
    db_session.commit()
    res = ds.to_dict()
    res.update({'vol_level': vol_level, 'vol_msg': vol_msg})
    return jsonify(res)


@app.route('/api/datasets/<int:dataset_id>/preview', methods=['GET'])
def preview_dataset(dataset_id):
    """获取数据预览（前 N 行 + 列类型）"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    rows = request.args.get('rows', 100, type=int)
    try:
        df = pd.read_excel(ds.file_path, sheet_name=ds.selected_sheet or 0,
                           engine='openpyxl', nrows=rows)
    except Exception as e:
        raise APIError(f'读取文件失败: {str(e)}')
    # 识别列类型
    col_types = {}
    for col in df.columns:
        col_str = str(col)
        if pd.api.types.is_numeric_dtype(df[col]):
            col_types[col_str] = 'numeric'
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            col_types[col_str] = 'datetime'
        else:
            col_types[col_str] = 'text'
    # 转前端友好格式
    preview_data = df.head(rows).to_dict(orient='records')
    # 处理 NaN
    for row in preview_data:
        for k, v in row.items():
            if pd.isna(v):
                row[k] = None
    return jsonify({
        'columns': [str(c) for c in df.columns],
        'col_types': col_types,
        'rows': preview_data,
        'total_rows': ds.row_count or len(df)
    })


@app.route('/api/datasets/<int:dataset_id>/preprocess', methods=['POST'])
def preprocess_dataset(dataset_id):
    """应用预处理选项并保存到数据库"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    data = request.get_json(silent=True) or {}
    import json
    ds.preprocessing_options = json.dumps(data, ensure_ascii=False)
    db_session.commit()
    # 返回预处理后的预览
    try:
        df = _get_dataframe(dataset_id)
        preview_data = df.head(100).to_dict(orient='records')
        for row in preview_data:
            for k, v in row.items():
                if pd.isna(v):
                    row[k] = None
        return jsonify({
            'columns': [str(c) for c in df.columns],
            'rows': preview_data,
            'total_rows': len(df),
            'message': '预处理已保存'
        })
    except Exception as e:
        raise APIError(f'预处理失败: {str(e)}')


@app.route('/api/datasets/<int:dataset_id>', methods=['DELETE'])
def delete_dataset(dataset_id):
    """删除数据集（同时删除关联图表）"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    # 删除关联图表
    db_session.query(Chart).filter_by(dataset_id=dataset_id).delete()
    # 删除文件
    try:
        if os.path.exists(ds.file_path):
            os.remove(ds.file_path)
    except Exception:
        pass
    db_session.delete(ds)
    db_session.commit()
    return jsonify({'message': '已删除'})


def _get_file_size_mb(filepath):
    try:
        return os.path.getsize(filepath) / (1024 * 1024)
    except Exception:
        return 0


@app.route('/api/datasets/<int:dataset_id>/info', methods=['GET'])
def get_dataset_info(dataset_id):
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
    size_mb = _get_file_size_mb(ds.file_path) if ds.file_path else 0
    return jsonify({
        'id': ds.id,
        'name': ds.name,
        'file_size_mb': round(size_mb, 2),
        'row_count': ds.row_count or 0,
        'column_count': ds.column_count or 0
    })


# -----------------------------------------------------------------------
# 趋势分析模块（一元线性回归 + 预测）
# -----------------------------------------------------------------------
from collections import namedtuple

TrendResult = namedtuple('TrendResult', ['slope', 'intercept', 'r2', 'direction',
                                         'x_pred', 'y_pred', 'x_fit', 'y_fit',
                                         'x_orig', 'y_orig'])

def trend_analysis(x, y):
    """对 x, y 进行一元线性回归分析，返回 TrendResult"""
    import numpy as np
    from sklearn.linear_model import LinearRegression
    from sklearn.metrics import r2_score

    x = np.array(x, dtype=float)
    y = np.array(y, dtype=float)

    # 清理 NaN / Inf
    mask = ~(np.isnan(x) | np.isnan(y) | np.isinf(x) | np.isinf(y))
    x = x[mask]
    y = y[mask]
    if len(x) < 3:
        return TrendResult(0, 0, 0, '平稳', [], [], [], [], x.tolist(), y.tolist())

    model = LinearRegression()
    model.fit(x.reshape(-1, 1), y)
    slope = float(model.coef_[0])
    intercept = float(model.intercept_)
    y_pred = model.predict(x.reshape(-1, 1))
    r2 = float(r2_score(y, y_pred))

    # 趋势方向
    if slope > 0.01:
        direction = '上升'
    elif slope < -0.01:
        direction = '下降'
    else:
        direction = '平稳'

    # 基于最后一步长外推未来 3 期
    if len(x) >= 2:
        step = (x[-1] - x[0]) / max(len(x) - 1, 1)
    else:
        step = 1
    x_pred = [x[-1] + step * (i + 1) for i in range(3)]
    y_pred_future = [float(model.predict([[xv]])[0]) for xv in x_pred]
    x_fit = x.tolist()
    y_fit = y_pred.tolist()

    return TrendResult(slope, intercept, r2, direction,
                       x_pred, y_pred_future, x_fit, y_fit,
                       x.tolist(), y.tolist())


def _build_plotly_json(df, chart, trend_result=None):
    """根据图表配置构建 Plotly 图形 JSON"""
    import plotly.graph_objects as go
    import json

    chart_config = json.loads(chart.config) if chart.config else {}
    title = chart_config.get('title', chart.name)
    x_label = chart_config.get('x_label', chart.x_col or '')
    y_label = chart_config.get('y_label', chart.y_col or '')
    color_theme = chart_config.get('color', '#1f77b4')
    chart_type = chart.chart_type or 'scatter'

    x_col = chart.x_col
    y_col = chart.y_col
    y2_col = chart.y2_col

    # 获取数据
    if x_col and x_col in df.columns and y_col and y_col in df.columns:
        x_data = df[x_col].dropna()
        y_data = df[y_col].dropna()
        # 对齐
        common = x_data.index.intersection(y_data.index)
        x_vals = x_data.loc[common].tolist()
        y_vals = y_data.loc[common].tolist()
    else:
        x_vals, y_vals = [], []

    fig = go.Figure()

    use_gl = len(x_vals) > 5000 and chart_type in ('scatter', 'line')

    # 主 trace
    if chart_type == 'scatter':
        trace_type = 'scattergl' if use_gl else 'scatter'
        fig.add_trace(go.Scattergl(
            x=x_vals, y=y_vals, mode='markers',
            name=chart.name, marker=dict(color=color_theme, size=6)
        ) if use_gl else go.Scatter(
            x=x_vals, y=y_vals, mode='markers',
            name=chart.name, marker=dict(color=color_theme, size=6)
        ))
    elif chart_type == 'line':
        trace_type = 'scattergl' if use_gl else 'scatter'
        fig.add_trace(go.Scattergl(
            x=x_vals, y=y_vals, mode='lines+markers',
            name=chart.name, marker=dict(color=color_theme),
            line=dict(color=color_theme)
        ) if use_gl else go.Scatter(
            x=x_vals, y=y_vals, mode='lines+markers',
            name=chart.name, marker=dict(color=color_theme),
            line=dict(color=color_theme)
        ))
    elif chart_type == 'bar':
        fig.add_trace(go.Bar(x=x_vals, y=y_vals, name=chart.name,
                             marker_color=color_theme))
    elif chart_type == 'area':
        fig.add_trace(go.Scatter(x=x_vals, y=y_vals, mode='lines',
                                 fill='tozeroy', name=chart.name,
                                 line=dict(color=color_theme)))
    elif chart_type == 'box':
        if y2_col and y2_col in df.columns:
            fig.add_trace(go.Box(y=df[y_col].dropna(), name=y_col,
                                 marker_color=color_theme))
            fig.add_trace(go.Box(y=df[y2_col].dropna(), name=y2_col,
                                 marker_color=color_theme))
        else:
            fig.add_trace(go.Box(y=y_vals, name=chart.name,
                                 marker_color=color_theme))
    elif chart_type == 'pie':
        if y2_col and y2_col in df.columns:
            labels = df[y2_col].dropna().unique()[:20]
            values = df.groupby(y2_col)[y_col].sum().values[:20]
        else:
            # 使用 X 列前 20 个值作为标签
            labels = x_vals[:20] if x_vals else []
            values = y_vals[:20] if y_vals else []
        fig.add_trace(go.Pie(labels=labels, values=values, name=chart.name))

    # 趋势线（散点/折线/柱状图）
    if trend_result and chart.trend_enabled and chart_type in ('scatter', 'line', 'bar'):
        # 拟合线
        fig.add_trace(go.Scatter(
            x=trend_result.x_fit, y=trend_result.y_fit,
            mode='lines', name='趋势线',
            line=dict(color='red', dash='dash', width=2)
        ))
        # 预测点
        fig.add_trace(go.Scatter(
            x=trend_result.x_pred, y=trend_result.y_pred,
            mode='markers', name='预测值',
            marker=dict(color='green', symbol='star', size=12)
        ))

    fig.update_layout(
        title=title,
        xaxis_title=x_label,
        yaxis_title=y_label,
        template='plotly_white',
        hovermode='closest',
        margin=dict(l=40, r=20, t=40, b=40),
        height=400
    )
    return fig


# -----------------------------------------------------------------------
# 图表管理
# -----------------------------------------------------------------------
@app.route('/api/spaces/<int:space_id>/charts', methods=['GET'])
def list_charts(space_id):
    """获取空间内所有图表"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    charts = db_session.query(Chart).filter_by(space_id=space_id).order_by(Chart.created_at.desc()).all()
    return jsonify([c.to_dict() for c in charts])


@app.route('/api/spaces/<int:space_id>/charts', methods=['POST'])
def create_chart(space_id):
    """创建图表"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    data = request.get_json(silent=True) or {}
    required = ['dataset_id', 'name', 'chart_type', 'x_col', 'y_col']
    for field in required:
        if field not in data:
            raise APIError(f'缺少必填字段: {field}')
    import json
    chart = Chart(
        space_id=space_id,
        dataset_id=data['dataset_id'],
        name=data['name'],
        chart_type=data.get('chart_type', 'scatter'),
        x_col=data.get('x_col', ''),
        y_col=data.get('y_col', ''),
        y2_col=data.get('y2_col', ''),
        trend_enabled=data.get('trend_enabled', True),
        config=json.dumps(data.get('config', {}), ensure_ascii=False)
    )
    db_session.add(chart)
    db_session.commit()
    return jsonify(chart.to_dict()), 201


@app.route('/api/charts/<int:chart_id>', methods=['PUT'])
def update_chart(chart_id):
    """更新图表"""
    chart = db_session.get(Chart, chart_id)
    if not chart:
        raise APIError('图表不存在', 404)
    data = request.get_json(silent=True) or {}
    import json
    if 'name' in data:
        chart.name = data['name']
    if 'chart_type' in data:
        chart.chart_type = data['chart_type']
    if 'x_col' in data:
        chart.x_col = data['x_col']
    if 'y_col' in data:
        chart.y_col = data['y_col']
    if 'y2_col' in data:
        chart.y2_col = data['y2_col']
    if 'trend_enabled' in data:
        chart.trend_enabled = data['trend_enabled']
    if 'config' in data:
        chart.config = json.dumps(data['config'], ensure_ascii=False)
    db_session.commit()
    return jsonify(chart.to_dict())


@app.route('/api/charts/<int:chart_id>', methods=['DELETE'])
def delete_chart(chart_id):
    """删除图表"""
    chart = db_session.get(Chart, chart_id)
    if not chart:
        raise APIError('图表不存在', 404)
    db_session.delete(chart)
    db_session.commit()
    return jsonify({'message': '已删除'})


@app.route('/api/charts/<int:chart_id>/render', methods=['GET'])
def render_chart(chart_id):
    """渲染图表（返回 Plotly HTML + 趋势分析信息）"""
    chart = db_session.get(Chart, chart_id)
    if not chart:
        raise APIError('图表不存在', 404)
    try:
        df = _get_dataframe(chart.dataset_id)
    except Exception as e:
        raise APIError(f'读取数据失败: {str(e)}')

    trend_result = None
    trend_info = None
    if chart.trend_enabled and chart.chart_type in ('scatter', 'line', 'bar'):
        x_col = chart.x_col
        y_col = chart.y_col
        if x_col and y_col and x_col in df.columns and y_col in df.columns:
            x_data = df[x_col].dropna()
            y_data = df[y_col].dropna()
            common = x_data.index.intersection(y_data.index)
            x_vals = x_data.loc[common].tolist()
            y_vals = y_data.loc[common].tolist()
            if len(x_vals) >= 3:
                try:
                    trend_result = trend_analysis(x_vals, y_vals)
                except Exception:
                    app.logger.warning(f'趋势分析失败: x_col={x_col}, y_col={y_col}')
                    trend_result = None
                if trend_result:
                    trend_info = {
                        'slope': round(trend_result.slope, 4),
                        'intercept': round(trend_result.intercept, 4),
                        'r2': round(trend_result.r2, 4),
                        'direction': trend_result.direction,
                    'predictions': [
                        {'period': i+1, 'x': round(trend_result.x_pred[i], 4),
                         'y_pred': round(trend_result.y_pred[i], 4)}
                        for i in range(3)
                    ],
                    'data_points': [
                        {'x': trend_result.x_orig[i], 'y': trend_result.y_orig[i],
                         'y_fit': trend_result.y_fit[i]}
                        for i in range(len(trend_result.x_orig))
                    ]
                }

    fig = _build_plotly_json(df, chart, trend_result)
    chart_html = fig.to_html(full_html=False, include_plotlyjs='cdn', div_id=f'chart-{chart_id}')

    return jsonify({
        'chart_html': chart_html,
        'chart_data': chart.to_dict(),
        'trend_info': trend_info
    })


@app.route('/api/charts/<int:chart_id>/export-csv', methods=['GET'])
def export_chart_csv(chart_id):
    """导出图表趋势分析 CSV"""
    chart = db_session.get(Chart, chart_id)
    if not chart:
        raise APIError('图表不存在', 404)
    try:
        df = _get_dataframe(chart.dataset_id)
    except Exception as e:
        raise APIError(f'读取数据失败: {str(e)}')
    x_col, y_col = chart.x_col, chart.y_col
    if not x_col or not y_col or x_col not in df.columns or y_col not in df.columns:
        raise APIError('列不存在')
    x_data = df[x_col].dropna()
    y_data = df[y_col].dropna()
    common = x_data.index.intersection(y_data.index)
    x_vals = x_data.loc[common].tolist()
    y_vals = y_data.loc[common].tolist()

    if len(x_vals) >= 3:
        try:
            tr = trend_analysis(x_vals, y_vals)
        except Exception:
            app.logger.warning(f'导出 CSV 趋势分析失败: x_col={x_col}, y_col={y_col}')
            tr = None
        if tr:
            out_df = pd.DataFrame({
                'X': tr.x_orig,
                'Y': tr.y_orig,
                'Y_fit': tr.y_fit
            })
            # 追加预测行
            pred_df = pd.DataFrame({
                'X': tr.x_pred,
                'Y': ['']*3,
                'Y_fit': tr.y_pred
            })
            out_df = pd.concat([out_df, pred_df], ignore_index=True)
        else:
            out_df = pd.DataFrame({'X': x_vals, 'Y': y_vals, 'Y_fit': ['']*len(x_vals)})
    else:
        out_df = pd.DataFrame({'X': x_vals, 'Y': y_vals, 'Y_fit': ['']*len(x_vals)})

    csv_path = os.path.join(UPLOAD_FOLDER, f'chart_{chart_id}_trend.csv')
    out_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    return send_file(csv_path, mimetype='text/csv',
                     as_attachment=True,
                     download_name=f'{chart.name}_趋势分析.csv')


@app.route('/api/charts/<int:chart_id>/export-image', methods=['GET', 'POST'])
def export_chart_image(chart_id):
    """导出图表为 PNG（使用 Plotly 的 kaleido 引擎）"""
    chart = db_session.get(Chart, chart_id)
    if not chart:
        raise APIError('图表不存在', 404)
    try:
        df = _get_dataframe(chart.dataset_id)
    except Exception as e:
        raise APIError(f'读取数据失败: {str(e)}')

    trend_result = None
    if chart.trend_enabled and chart.chart_type in ('scatter', 'line', 'bar'):
        x_col, y_col = chart.x_col, chart.y_col
        if x_col and y_col and x_col in df.columns and y_col in df.columns:
            x_data = df[x_col].dropna()
            y_data = df[y_col].dropna()
            common = x_data.index.intersection(y_data.index)
            x_vals = x_data.loc[common].tolist()
            y_vals = y_data.loc[common].tolist()
            if len(x_vals) >= 3:
                try:
                    trend_result = trend_analysis(x_vals, y_vals)
                except Exception:
                    app.logger.warning(f'导出图片趋势分析失败: x_col={x_col}, y_col={y_col}')
                    trend_result = None

    fig = _build_plotly_json(df, chart, trend_result)
    img_path = os.path.join(UPLOAD_FOLDER, f'chart_{chart_id}.png')
    fig.write_image(img_path, format='png', width=900, height=500, scale=2)
    return send_file(img_path, mimetype='image/png',
                     as_attachment=True,
                     download_name=f'{chart.name}.png')


# -----------------------------------------------------------------------
# AI 配置管理
# -----------------------------------------------------------------------
@app.route('/api/ai-configs', methods=['GET'])
def list_ai_configs():
    """获取所有 AI 配置"""
    configs = db_session.query(AIConfig).all()
    return jsonify([c.to_dict() for c in configs])


@app.route('/api/ai-configs', methods=['POST'])
def create_ai_config():
    """新增 AI 配置"""
    data = request.get_json(silent=True) or {}
    required = ['name', 'base_url', 'api_key', 'model']
    for field in required:
        if not data.get(field):
            raise APIError(f'缺少必填字段: {field}')
    config = AIConfig(
        name=data['name'],
        base_url=data['base_url'].rstrip('/'),
        api_key=data['api_key'],
        model=data['model'],
        system_prompt=data.get('system_prompt', '你是一个数据分析助手，基于用户上传的 Excel 数据回答问题。'),
        max_tokens=data.get('max_tokens', 2000),
        temperature=data.get('temperature', 0.7),
        is_default=data.get('is_default', False)
    )
    if config.is_default:
        db_session.query(AIConfig).filter(AIConfig.is_default == True).update({'is_default': False})
    db_session.add(config)
    db_session.commit()
    return jsonify(config.to_dict()), 201


@app.route('/api/ai-configs/<int:config_id>', methods=['PUT'])
def update_ai_config(config_id):
    """更新 AI 配置"""
    config = db_session.get(AIConfig, config_id)
    if not config:
        raise APIError('AI 配置不存在', 404)
    data = request.get_json(silent=True) or {}
    for field in ['name', 'base_url', 'api_key', 'model', 'system_prompt',
                  'max_tokens', 'temperature', 'is_default']:
        if field in data:
            setattr(config, field, data[field])
    if data.get('is_default'):
        db_session.query(AIConfig).filter(AIConfig.is_default == True, AIConfig.id != config_id).update({'is_default': False})
    db_session.commit()
    return jsonify(config.to_dict())


@app.route('/api/ai-configs/<int:config_id>', methods=['DELETE'])
def delete_ai_config(config_id):
    """删除 AI 配置"""
    config = db_session.get(AIConfig, config_id)
    if not config:
        raise APIError('AI 配置不存在', 404)
    db_session.delete(config)
    db_session.commit()
    return jsonify({'message': '已删除'})


@app.route('/api/ai-configs/<int:config_id>/set-default', methods=['POST'])
def set_default_ai_config(config_id):
    """设置默认 AI 配置"""
    config = db_session.get(AIConfig, config_id)
    if not config:
        raise APIError('AI 配置不存在', 404)
    db_session.query(AIConfig).filter(AIConfig.is_default == True).update({'is_default': False})
    config.is_default = True
    db_session.commit()
    return jsonify(config.to_dict())


@app.route('/api/ai-configs/<int:config_id>/refresh-models', methods=['POST'])
def refresh_ai_models(config_id):
    config = db_session.get(AIConfig, config_id)
    if not config:
        raise APIError('AI 配置不存在', 404)
    import requests as http_requests
    models, _ = _fetch_models_from_url(config.base_url, config.api_key)
    if not models:
        raise APIError(f'获取模型列表失败，请检查 base_url 和 api_key 是否正确')
    config.cached_models = json.dumps(models, ensure_ascii=False)
    db_session.commit()
    return jsonify({'models': models, 'message': f'已刷新，获取到 {len(models)} 个模型'})


@app.route('/api/ai-configs/refresh-models-preview', methods=['POST'])
def refresh_ai_models_preview():
    """预览刷新模型（无需保存配置，直接使用表单传参）"""
    data = request.get_json(silent=True) or {}
    base_url = (data.get('base_url') or '').strip()
    api_key = (data.get('api_key') or '').strip()
    if not base_url or not api_key:
        raise APIError('base_url 和 api_key 不能为空')
    import requests as http_requests
    models, last_error = _fetch_models_from_url(base_url, api_key)
    if not models:
        raise APIError(f'获取模型列表失败: {last_error or "未知错误"}')
    return jsonify({'models': models, 'message': f'获取到 {len(models)} 个模型'})


def _fetch_models_from_url(base_url, api_key):
    """从 API 地址拉取模型列表，返回 (models_list, last_error_str)"""
    import requests as http_requests
    base = base_url.rstrip('/')
    candidates = []
    if '/v1' not in base:
        candidates.append(f"{base}/v1/models")
    candidates.append(f"{base}/models")
    last_error = ''
    for url in candidates:
        try:
            resp = http_requests.get(
                url,
                headers={'Authorization': f'Bearer {api_key}'},
                timeout=15
            )
            if resp.status_code == 200:
                data = resp.json()
                models = []
                if 'data' in data:
                    for m in data['data']:
                        if isinstance(m, dict) and 'id' in m:
                            models.append(m['id'])
                        elif isinstance(m, str):
                            models.append(m)
                if models:
                    return models, ''
                last_error = '返回数据中没有模型'
            else:
                last_error = f'HTTP {resp.status_code}'
        except requests.exceptions.RequestException as ex:
            last_error = str(ex)
            continue
    return [], last_error


# -----------------------------------------------------------------------
# 空间 AI 绑定
# -----------------------------------------------------------------------
@app.route('/api/spaces/<int:space_id>/ai-config', methods=['GET'])
def get_space_ai_config(space_id):
    """获取空间绑定的 AI 配置 ID"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    binding = db_session.query(SpaceAIConfig).filter_by(space_id=space_id).first()
    if binding:
        return jsonify({'ai_config_id': binding.ai_config_id})
    # 返回默认配置
    default = db_session.query(AIConfig).filter_by(is_default=True).first()
    return jsonify({'ai_config_id': default.id if default else None})


@app.route('/api/spaces/<int:space_id>/ai-config', methods=['POST'])
def set_space_ai_config(space_id):
    """设置空间绑定的 AI 配置"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    data = request.get_json(silent=True) or {}
    ai_config_id = data.get('ai_config_id')
    if not ai_config_id:
        raise APIError('请指定 AI 配置')
    ai_config = db_session.get(AIConfig, ai_config_id)
    if not ai_config:
        raise APIError('AI 配置不存在')
    # 删除旧绑定
    db_session.query(SpaceAIConfig).filter_by(space_id=space_id).delete()
    binding = SpaceAIConfig(space_id=space_id, ai_config_id=ai_config_id)
    db_session.add(binding)
    db_session.commit()
    return jsonify(binding.to_dict())


# -----------------------------------------------------------------------
# AI 聊天
# -----------------------------------------------------------------------
@app.route('/api/chat', methods=['POST'])
def chat():
    """发送消息给 AI（含数据集上下文）"""
    data = request.get_json(silent=True) or {}
    space_id = data.get('space_id')
    if not space_id:
        raise APIError('缺少 space_id')
    message = data.get('message', '').strip()
    if not message:
        raise APIError('消息不能为空')
    dataset_id = data.get('dataset_id')

    # 获取空间绑定的 AI 配置
    binding = db_session.query(SpaceAIConfig).filter_by(space_id=space_id).first()
    config = None
    if binding:
        config = db_session.get(AIConfig, binding.ai_config_id)
    if not config:
        config = db_session.query(AIConfig).filter_by(is_default=True).first()
    if not config:
        raise APIError('未配置 AI，请先在 AI 配置管理中添加配置并与空间绑定')

    # 存储用户消息
    user_msg = ChatHistory(space_id=space_id, role='user', content=message, dataset_id=dataset_id)
    db_session.add(user_msg)
    db_session.commit()

    # 构建 system prompt
    system_prompt = config.system_prompt or '你是一个数据分析助手。'
    if dataset_id:
        try:
            ds = db_session.get(Dataset, dataset_id)
            if ds:
                df = _get_dataframe(dataset_id)
                col_info = ', '.join([f'{c}({str(df[c].dtype)})' for c in df.columns])
                preview = df.head(10).to_string()
                desc = df.describe().to_string() if len(df.select_dtypes(include=[np.number]).columns) > 0 else '无数值列'
                data_context = (
                    f'\n\n## 当前数据集: {ds.name}\n'
                    f'- 列信息: {col_info}\n'
                    f'- 行数: {len(df)}\n'
                    f'- 前 10 行数据:\n{preview}\n'
                    f'- 描述性统计:\n{desc}\n'
                )
                system_prompt += data_context
        except Exception as e:
            app.logger.error(f'读取数据集失败: {str(e)}')

    # 构建消息列表
    messages = [{'role': 'system', 'content': system_prompt}]
    recent_history = db_session.query(ChatHistory).filter_by(space_id=space_id)\
        .order_by(ChatHistory.created_at.desc()).limit(20).all()
    recent_history.reverse()
    for h in recent_history:
        messages.append({'role': h.role, 'content': h.content})
    # 确保最后一条是 user 消息
    if not messages or messages[-1]['role'] != 'user':
        messages.append({'role': 'user', 'content': message})

    # 调用 AI API
    import requests as http_requests
    try:
        resp = http_requests.post(
            f"{config.base_url}/chat/completions",
            headers={
                'Authorization': f'Bearer {config.api_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': config.model,
                'messages': messages,
                'max_tokens': config.max_tokens or 2000,
                'temperature': config.temperature or 0.7
            },
            timeout=60
        )
        if resp.status_code != 200:
            raise APIError(f'AI 接口返回错误: {resp.status_code} - {resp.text[:200]}')
        reply = resp.json()['choices'][0]['message']['content']
    except Exception as e:
        if isinstance(e, APIError):
            raise
        raise APIError(f'调用 AI 接口失败: {str(e)}')

    # 存储 AI 回复
    ai_msg = ChatHistory(space_id=space_id, role='assistant', content=reply, dataset_id=dataset_id)
    db_session.add(ai_msg)
    db_session.commit()

    return jsonify({'reply': reply, 'message_id': ai_msg.id})


@app.route('/api/spaces/<int:space_id>/chat-history', methods=['GET'])
def get_chat_history(space_id):
    """获取空间聊天记录"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    history = db_session.query(ChatHistory).filter_by(space_id=space_id)\
        .order_by(ChatHistory.created_at.asc()).all()
    return jsonify([h.to_dict() for h in history])


@app.route('/api/spaces/<int:space_id>/chat-history', methods=['DELETE'])
def clear_chat_history(space_id):
    """清空空间聊天记录"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    db_session.query(ChatHistory).filter_by(space_id=space_id).delete()
    db_session.commit()
    return jsonify({'message': '已清空'})


# -----------------------------------------------------------------------
# 分析笔记
# -----------------------------------------------------------------------
@app.route('/api/notes', methods=['POST'])
def create_note():
    """保存分析笔记"""
    data = request.get_json(silent=True) or {}
    space_id = data.get('space_id')
    if not space_id:
        raise APIError('缺少 space_id')
    note = AnalysisNote(
        space_id=space_id,
        title=data.get('title', '分析笔记'),
        content=data.get('content', '')
    )
    db_session.add(note)
    db_session.commit()
    return jsonify(note.to_dict()), 201


@app.route('/api/spaces/<int:space_id>/notes', methods=['GET'])
def list_notes(space_id):
    """获取空间笔记列表"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
    notes = db_session.query(AnalysisNote).filter_by(space_id=space_id)\
        .order_by(AnalysisNote.created_at.desc()).all()
    return jsonify([n.to_dict() for n in notes])


@app.route('/api/notes/<int:note_id>', methods=['DELETE'])
def delete_note(note_id):
    """删除分析笔记"""
    note = db_session.get(AnalysisNote, note_id)
    if not note:
        raise APIError('笔记不存在', 404)
    db_session.delete(note)
    db_session.commit()
    return jsonify({'message': '已删除'})


# -----------------------------------------------------------------------
# 导出与备份
# -----------------------------------------------------------------------
@app.route('/api/spaces/<int:space_id>/export', methods=['GET'])
def export_space(space_id):
    """导出整个空间为 ZIP 包"""
    import zipfile
    import json as json_module
    from io import BytesIO

    space = db_session.get(Space, space_id)
    if not space:
        raise APIError('空间不存在', 404)

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 数据集文件
        datasets = db_session.query(Dataset).filter_by(space_id=space_id).all()
        for ds in datasets:
            if os.path.exists(ds.file_path):
                arc_name = f'datasets/{ds.name}_{ds.id}{os.path.splitext(ds.file_path)[1]}'
                zf.write(ds.file_path, arc_name)
        # 图表 HTML 快照
        charts = db_session.query(Chart).filter_by(space_id=space_id).all()
        for chart in charts:
            try:
                url = f"http://127.0.0.1:{os.environ.get('PORT', 5000)}/api/charts/{chart.id}/render"
                import requests as http_requests
                resp = http_requests.get(url, timeout=10)
                if resp.status_code == 200:
                    zf.writestr(f'charts/{chart.name}_{chart.id}.json',
                                json_module.dumps(resp.json(), ensure_ascii=False, indent=2))
            except Exception:
                pass
        # 聊天记录
        history = db_session.query(ChatHistory).filter_by(space_id=space_id)\
            .order_by(ChatHistory.created_at.asc()).all()
        zf.writestr('chat_history.json',
                    json_module.dumps([h.to_dict() for h in history],
                                      ensure_ascii=False, indent=2,
                                      default=str))
        # 分析笔记
        notes = db_session.query(AnalysisNote).filter_by(space_id=space_id).all()
        zf.writestr('analysis_notes.json',
                    json_module.dumps([n.to_dict() for n in notes],
                                      ensure_ascii=False, indent=2,
                                      default=str))
    buf.seek(0)
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True,
                     download_name=f'{space.name}_导出.zip')


@app.route('/api/system/backup', methods=['GET'])
def backup_database():
    """下载数据库备份"""
    if not os.path.exists(DB_PATH):
        raise APIError('数据库文件不存在')
    return send_file(DB_PATH, mimetype='application/octet-stream',
                     as_attachment=True,
                     download_name=f'excelany_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')


@app.route('/api/system/restore', methods=['POST'])
def restore_database():
    """上传备份文件恢复数据库"""
    if 'file' not in request.files:
        raise APIError('请上传备份文件')
    file = request.files['file']
    if file.filename == '':
        raise APIError('请选择文件')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext != '.db':
        raise APIError('请上传 .db 文件')
    # 备份当前数据库
    backup_path = os.path.join(BACKUP_FOLDER, f'pre_restore_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')
    if os.path.exists(DB_PATH):
        import shutil
        shutil.copy2(DB_PATH, backup_path)
    # 关闭当前连接并覆盖
    db_session.remove()
    file.save(DB_PATH)
    # 重新初始化
    init_db()
    return jsonify({'message': '数据库已恢复，原数据库已备份到 backup 目录'})


@app.route('/api/system/shutdown', methods=['POST'])
def shutdown_server():
    import signal
    app.logger.info('收到退出指令，正在关闭服务...')
    # 使用线程延迟执行，确保 HTTP 响应能返回
    def delayed_shutdown():
        import time
        time.sleep(0.5)
        # 发送信号让主线程退出
        os.kill(os.getpid(), signal.SIGTERM)
    threading.Thread(target=delayed_shutdown, daemon=True).start()
    return jsonify({'message': '服务正在关闭...'})


@app.route('/')
def index():
    """返回前端 SPA 页面"""
    return render_template('index.html')

# ============================== 启动入口 =====================================

def init_db():
    """初始化数据库，创建所有表，并同步缺失的列"""
    Base.metadata.create_all(bind=engine)
    
    # 简单的数据库迁移逻辑：检查并添加缺失的列
    inspector = inspect(engine)
    
    # 检查 ai_config 表
    if 'ai_config' in inspector.get_table_names():
        columns = [c['name'] for c in inspector.get_columns('ai_config')]
        if 'cached_models' not in columns:
            app.logger.info("正在同步数据库：为 ai_config 表添加 cached_models 列")
            try:
                with engine.connect() as conn:
                    conn.execute(text("ALTER TABLE ai_config ADD COLUMN cached_models TEXT DEFAULT '[]'"))
                    conn.commit()
            except Exception as e:
                app.logger.error(f"同步 ai_config 表失败: {e}")

def main():
    parser = argparse.ArgumentParser(description='ExcelAny Backend Server')
    parser.add_argument('--port', type=int, default=5000, help='服务运行端口')
    parser.add_argument('--password', type=str, help='访问密码')
    args = parser.parse_args()

    global ACCESS_PASSWORD
    ACCESS_PASSWORD = args.password

    init_db()
    port = args.port
    debug = not getattr(sys, 'frozen', False)
    
    # 注册信号处理，支持优雅退出
    import signal
    def handle_exit(signum, frame):
        app.logger.info(f'收到信号 {signum}，正在退出...')
        sys.exit(0)
    signal.signal(signal.SIGTERM, handle_exit)
    signal.signal(signal.SIGINT, handle_exit)
    
    # 打包环境自动打开浏览器
    if not debug:
        import webbrowser
        
        def open_browser():
            import time
            time.sleep(1.5)
            webbrowser.open(f'http://127.0.0.1:{port}')
            
        threading.Thread(target=open_browser, daemon=True).start()
        
    app.logger.info(f'服务启动: http://127.0.0.1:{port}')
    if ACCESS_PASSWORD:
        app.logger.info('已开启启动密码验证')
    
    app.run(host='127.0.0.1', port=port, debug=debug, use_reloader=False)

if __name__ == '__main__':
    main()