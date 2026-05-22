import os
import uuid
import pandas as pd
import numpy as np
from flask import Blueprint, jsonify, request
from core.database import db_session, UPLOAD_FOLDER
from core.exceptions import APIError
from models.models import Space, Dataset, Chart
from services.preprocess_service import get_dataframe, get_file_size_mb

datasets_bp = Blueprint('datasets', __name__)

def categorize_data_volume(row_count):
    """数据量级判定与提醒信息"""
    if row_count >= 1000000:
        return "million", f"该文件包含 {row_count} 行数据，属于百万行级及以上海量表格。加载、渲染、运算耗时较长，建议开启采样以保证流畅度。"
    elif row_count >= 100000:
        return "hundred_thousand", f"该文件包含 {row_count} 行数据，属于十万行级大文件，加载预计耗时较长。"
    elif row_count >= 10000:
        return "ten_thousand", f"该文件包含 {row_count} 行数据，属于万行级数据，加载较快。"
    return "small", f"该文件包含 {row_count} 行数据，属于常规量级。"


def get_row_count(file_path, sheet_name=None):
    """高效获取 Excel 行数（通过只读模式载入 workbook，节约内存）"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, keep_links=False)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        count = ws.max_row
        wb.close()
        return count
    except Exception:
        # 回退机制：使用 pandas 读取首列计算行数
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, usecols=[0], engine='openpyxl')
            return len(df)
        except Exception:
            return 0


@datasets_bp.route('/api/spaces/<int:space_id>/datasets', methods=['POST'])
def upload_dataset(space_id):
    """上传 Excel 数据集（第一步：接收物理文件并返回工作表列表）"""
    check_space = db_session.get(Space, space_id)
    if not check_space:
        raise APIError('空间不存在', 404)
        
    if 'file' not in request.files:
        raise APIError('请上传文件')
    file = request.files['file']
    if file.filename == '':
        raise APIError('请选择文件')
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        raise APIError('仅支持 .xlsx 或 .xls 文件')
        
    # 重命名物理文件并落盘
    safe_name = f"{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(UPLOAD_FOLDER, safe_name)
    file.save(save_path)
    
    # 提取工作表列表
    try:
        xls = pd.ExcelFile(save_path, engine='openpyxl')
        sheet_names = xls.sheet_names
        xls.close()
    except Exception as e:
        if os.path.exists(save_path):
            os.remove(save_path)
        raise APIError(f'解析 Excel 失败: {str(e)}')

    ds = Dataset(
        space_id=space_id,
        name=os.path.splitext(file.filename)[0],
        file_path=save_path,
        selected_sheet=None,
        row_count=None
    )
    db_session.add(ds)
    db_session.commit()
    
    # 获取大致行数并判定数据级
    ds.row_count = get_row_count(save_path)
    db_session.commit()
    
    vol_level, vol_msg = categorize_data_volume(ds.row_count)
    return jsonify({
        'dataset_id': ds.id, 
        'sheets': sheet_names, 
        'name': ds.name,
        'vol_level': vol_level,
        'vol_msg': vol_msg
    })


@datasets_bp.route('/api/datasets/<int:dataset_id>/sheets', methods=['GET'])
def get_dataset_sheets(dataset_id):
    """获取工作簿的所有工作表名称"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    try:
        xls = pd.ExcelFile(ds.file_path, engine='openpyxl')
        sheets = xls.sheet_names
        xls.close()
        return jsonify(sheets)
    except Exception as e:
        raise APIError(f'读取工作表失败: {str(e)}')


@datasets_bp.route('/api/datasets/<int:dataset_id>/confirm-sheet', methods=['POST'])
def confirm_dataset_sheet(dataset_id):
    """确认选择的工作表并更新行数"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    data = request.get_json(silent=True) or {}
    sheet = data.get('sheet', '')
    if sheet:
        ds.selected_sheet = sheet
        
    ds.row_count = get_row_count(ds.file_path, sheet)
    vol_level, vol_msg = categorize_data_volume(ds.row_count or 0)
    db_session.commit()
    
    res = ds.to_dict()
    res.update({'vol_level': vol_level, 'vol_msg': vol_msg})
    return jsonify(res)


@datasets_bp.route('/api/datasets/<int:dataset_id>/preview', methods=['GET'])
def preview_dataset(dataset_id):
    """获取数据集预览（限制前 N 行）"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    rows = request.args.get('rows', 100, type=int)
    try:
        df = pd.read_excel(ds.file_path, sheet_name=ds.selected_sheet or 0,
                           engine='openpyxl', nrows=rows)
    except Exception as e:
        raise APIError(f'读取数据失败: {str(e)}')
        
    col_types = {}
    for col in df.columns:
        col_str = str(col)
        if pd.api.types.is_numeric_dtype(df[col]):
            col_types[col_str] = 'numeric'
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            col_types[col_str] = 'datetime'
        else:
            col_types[col_str] = 'text'
            
    preview_data = df.head(rows).to_dict(orient='records')
    for row in preview_data:
        for k, v in row.items():
            if pd.isna(v):
                row[k] = None
                
    return jsonify({
        'columns': [str(c) for c in df.columns],
        'col_types': col_types,
        'rows': preview_data,
        'total_rows': ds.row_count or len(df)
    })


@datasets_bp.route('/api/datasets/<int:dataset_id>/preprocess', methods=['POST'])
def preprocess_dataset(dataset_id):
    """应用并保存预处理与采样选项参数"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    data = request.get_json(silent=True) or {}
    import json
    ds.preprocessing_options = json.dumps(data, ensure_ascii=False)
    db_session.commit()
    
    try:
        df = get_dataframe(dataset_id)
        preview_data = df.head(100).to_dict(orient='records')
        for row in preview_data:
            for k, v in row.items():
                if pd.isna(v):
                    row[k] = None
        return jsonify({
            'columns': [str(c) for c in df.columns],
            'rows': preview_data,
            'total_rows': len(df),
            'message': '预处理已保存'
        })
    except Exception as e:
        raise APIError(f'数据策略处理失败: {str(e)}')


@datasets_bp.route('/api/datasets/<int:dataset_id>', methods=['PUT'])
def rename_dataset(dataset_id):
    """修改数据集的友好展示名称"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    if not name or not name.strip():
        raise APIError('数据集名称不能为空')
        
    ds.name = name.strip()
    db_session.commit()
    return jsonify(ds.to_dict())


@datasets_bp.route('/api/datasets/<int:dataset_id>', methods=['DELETE'])
def delete_dataset(dataset_id):
    """物理删除数据集与绑定的图表"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    db_session.query(Chart).filter_by(dataset_id=dataset_id).delete()
    try:
        if os.path.exists(ds.file_path):
            os.remove(ds.file_path)
    except Exception:
        pass
        
    db_session.delete(ds)
    db_session.commit()
    return jsonify({'message': '已删除'})


@datasets_bp.route('/api/datasets/<int:dataset_id>/info', methods=['GET'])
def get_dataset_info(dataset_id):
    """获取数据集大小、行数与存储列的简要信息"""
    ds = db_session.get(Dataset, dataset_id)
    if not ds:
        raise APIError('数据集不存在', 404)
        
    size_mb = get_file_size_mb(ds.file_path) if ds.file_path else 0
    return jsonify({
        'id': ds.id,
        'name': ds.name,
        'file_size_mb': round(size_mb, 2),
        'row_count': ds.row_count or 0
    })
